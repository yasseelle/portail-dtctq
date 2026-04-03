import sys
import os
from pathlib import Path
from datetime import datetime

# ── Add backend to path ───────────────────────────────────────────────────────
sys.path.insert(0, str(Path(__file__).parent))

from database import SessionLocal, engine, Base
import models

Base.metadata.create_all(bind=engine)
db = SessionLocal()

# ── CONFIG ────────────────────────────────────────────────────────────────────
COURRIER_EXCEL        = r"C:\courrier\etat_courrier_2026.xlsx"
BORDEREAU_EXCEL       = r"C:\bordereau_envoi\etat_bordereaux_envoi_2026.xlsx"
DEPART_EXCEL          = r"C:\courrier_depart_reception\etat_courrier_2026.xlsx"

DATA_START_ROW = 6  # row index (1-based) where data starts

MONTHS_MAP = {
    "janvier":1,"février":2,"fevrier":2,"mars":3,"avril":4,
    "mai":5,"mai":5,"juin":6,"juillet":7,"août":8,"aout":8,
    "septembre":9,"octobre":10,"novombr":11,"novembre":11,"december":12,"décembre":12,
}

def parse_date(val):
    if not val: return None
    s = str(val).strip()
    for fmt in ("%d/%m/%Y","%Y-%m-%d","%d-%m-%Y"):
        try: return datetime.strptime(s, fmt).strftime("%d/%m/%Y")
        except: pass
    return s

def clean(val):
    if val is None: return ""
    return str(val).strip()

# =============================================================================
# ── 1. COURRIER ARRIVÉE ───────────────────────────────────────────────────────
# =============================================================================
def import_courrier():
    try:
        import openpyxl
        wb = openpyxl.load_workbook(COURRIER_EXCEL, read_only=True)
    except FileNotFoundError:
        print(f"  ⚠️  Fichier introuvable : {COURRIER_EXCEL}")
        return 0

    created = skipped = 0
    for sheet_name in wb.sheetnames:
        month_key = sheet_name.lower().strip()
        month_num = MONTHS_MAP.get(month_key, 0)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        for row in rows[DATA_START_ROW - 1:]:
            if not any(v for v in row): continue

            num        = row[0]
            expediteur = clean(row[1]) if len(row) > 1 else ""
            date_val   = parse_date(row[2]) if len(row) > 2 else ""
            objet      = clean(row[3]) if len(row) > 3 else ""

            if not expediteur and not objet: continue

            # Duplicate check
            exists = db.query(models.Courrier).filter(
                models.Courrier.expediteur   == expediteur,
                models.Courrier.date_courrier == date_val,
                models.Courrier.objet        == objet,
                models.Courrier.mois         == sheet_name,
            ).first()

            if exists:
                skipped += 1
                continue

            record = models.Courrier(
                expediteur    = expediteur,
                date_courrier = date_val,
                objet         = objet,
                pdf_path      = "",
                txt_path      = "",
                mois          = sheet_name,
            )
            db.add(record)
            created += 1

    db.commit()
    wb.close()
    print(f"  ✅ Courrier Arrivée  : {created} créés, {skipped} ignorés (doublons)")
    return created

# =============================================================================
# ── 2. BORDEREAU D'ENVOI ──────────────────────────────────────────────────────
# =============================================================================
def import_bordereau():
    try:
        import openpyxl
        wb = openpyxl.load_workbook(BORDEREAU_EXCEL, read_only=True)
    except FileNotFoundError:
        print(f"  ⚠️  Fichier introuvable : {BORDEREAU_EXCEL}")
        return 0

    created = skipped = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        for row in rows[DATA_START_ROW - 1:]:
            if not any(v for v in row): continue

            reference    = clean(row[1]) if len(row) > 1 else ""
            date_val     = parse_date(row[2]) if len(row) > 2 else ""
            destinataire = clean(row[3]) if len(row) > 3 else ""
            objet        = clean(row[4]) if len(row) > 4 else ""

            if not reference and not objet: continue

            # Duplicate check
            exists = db.query(models.Bordereau).filter(
                models.Bordereau.reference == reference,
                models.Bordereau.objet     == objet,
            ).first()

            if exists:
                skipped += 1
                continue

            record = models.Bordereau(
                reference    = reference,
                destinataire = destinataire,
                objet        = objet,
                pdf_path     = "",
                txt_path     = "",
            )
            db.add(record)
            created += 1

    db.commit()
    wb.close()
    print(f"  ✅ Bordereau Envoi   : {created} créés, {skipped} ignorés (doublons)")
    return created

# =============================================================================
# ── 3. COURRIER DÉPART / RÉCEPTION ───────────────────────────────────────────
# =============================================================================
def import_depart_reception():
    try:
        import openpyxl
        wb = openpyxl.load_workbook(DEPART_EXCEL, read_only=True)
    except FileNotFoundError:
        print(f"  ⚠️  Fichier introuvable : {DEPART_EXCEL}")
        return 0

    created = skipped = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        for row in rows[DATA_START_ROW - 1:]:
            if not any(v for v in row): continue

            reference      = clean(row[1]) if len(row) > 1 else ""
            date_depart    = parse_date(row[2]) if len(row) > 2 else ""
            destinataire   = clean(row[3]) if len(row) > 3 else ""
            objet          = clean(row[4]) if len(row) > 4 else ""
            date_reception = parse_date(row[7]) if len(row) > 7 else ""

            if not reference and not objet: continue

            # Duplicate check
            exists = db.query(models.CourrierDepart).filter(
                models.CourrierDepart.reference == reference,
            ).first()

            if exists:
                skipped += 1
                continue

            record = models.CourrierDepart(
                reference        = reference,
                date_depart      = date_depart,
                destinataire     = destinataire,
                objet            = objet,
                pdf_depart_path  = "",
                pdf_reception_path = "",
                date_reception   = date_reception,
                mois             = sheet_name,
            )
            db.add(record)
            created += 1

    db.commit()
    wb.close()
    print(f"  ✅ Courrier Départ/Réc: {created} créés, {skipped} ignorés (doublons)")
    return created


# =============================================================================
# ── MAIN ─────────────────────────────────────────────────────────────────────
# =============================================================================
if __name__ == "__main__":
    print("\n" + "="*55)
    print("  IMPORT EXCEL → BASE DE DONNÉES")
    print("="*55)

    c1 = import_courrier()
    c2 = import_bordereau()
    c3 = import_depart_reception()

    print(f"\n  📊 Total importé : {c1 + c2 + c3} enregistrements")
    print("="*55 + "\n")

    db.close()
