"""
excel_attachements.py — Sync DB → Excel pour les Attachements ONEE
Placé dans : C:\projets\portail-dtctq\backend\excel_attachements.py
"""

import re
from pathlib import Path
from datetime import datetime
from typing import Optional

from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

from models import Attachement

# ── Chemin du fichier Excel ───────────────────────────────────────────────────
EXCEL_PATH = Path(r"C:\projets\portail-dtctq\backend\etat_attachements.xlsx")

# ── Couleurs ONEE (vert + blanc) ──────────────────────────────────────────────
COLOR_HEADER_BG  = "1B5E20"   # vert foncé
COLOR_HEADER_FG  = "FFFFFF"   # blanc
COLOR_ROW_ODD    = "F1F8E9"   # vert très clair
COLOR_ROW_EVEN   = "FFFFFF"   # blanc
COLOR_BORDER     = "A5D6A7"   # vert moyen
COLOR_ACCENT     = "2E7D32"   # vert moyen-foncé (liens)
COLOR_SUBHEADER  = "C8E6C9"   # vert pâle pour sous-titres

# ── Colonnes principales ──────────────────────────────────────────────────────
MAIN_COLUMNS = [
    ("Entreprise",       22),
    ("Date Document",    16),
    ("Marché N°",        14),
    ("Marché Nom",       35),
    ("Date Début",       14),
    ("Date Fin",         14),
    ("Att N°",           9),
    ("Nbre Articles",    14),
    ("Montant Total DH", 18),
    ("Source OCR",       14),
    ("PDF",              40),
]

# ── Helpers de style ──────────────────────────────────────────────────────────
def _thin_border(color: str = "BDBDBD") -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)

def _header_font()    -> Font:  return Font(bold=True, color=COLOR_HEADER_FG, name="Calibri", size=11)
def _normal_font()    -> Font:  return Font(name="Calibri", size=10)
def _link_font()      -> Font:  return Font(name="Calibri", size=10, color=COLOR_ACCENT, underline="single")
def _centered()       -> Alignment: return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _left_aligned()   -> Alignment: return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=COLOR_HEADER_BG)

def _row_fill(idx: int) -> PatternFill:
    color = COLOR_ROW_ODD if idx % 2 == 0 else COLOR_ROW_EVEN
    return PatternFill("solid", fgColor=color)


# ── Création / mise à jour du fichier Excel ───────────────────────────────────
def sync_to_excel(db: Session) -> None:
    """
    Recrée complètement l'onglet 'Attachements' dans etat_attachements.xlsx.
    Préserve les autres onglets s'ils existent.
    """
    # Charge ou crée le workbook
    if EXCEL_PATH.exists():
        wb = load_workbook(EXCEL_PATH)
    else:
        wb = Workbook()
        # Supprime la feuille vide par défaut
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # Supprime l'ancienne feuille attachements
    if "Attachements" in wb.sheetnames:
        del wb["Attachements"]

    ws = wb.create_sheet("Attachements", 0)

    # ── Titre principal ───────────────────────────────────────────────────────
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value     = "📎 REGISTRE DES ATTACHEMENTS — ONEE DTC/TQ"
    title_cell.font      = Font(bold=True, size=14, color=COLOR_HEADER_FG, name="Calibri")
    title_cell.fill      = PatternFill("solid", fgColor=COLOR_ACCENT)
    title_cell.alignment = _centered()
    ws.row_dimensions[1].height = 30

    # ── Date de mise à jour ───────────────────────────────────────────────────
    ws.merge_cells("A2:K2")
    date_cell = ws["A2"]
    date_cell.value     = f"Mis à jour le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    date_cell.font      = Font(italic=True, size=9, color="616161", name="Calibri")
    date_cell.alignment = _centered()
    ws.row_dimensions[2].height = 18

    # ── En-têtes colonnes (ligne 3) ───────────────────────────────────────────
    ws.row_dimensions[3].height = 28
    for col_idx, (col_name, col_width) in enumerate(MAIN_COLUMNS, start=1):
        cell             = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font        = _header_font()
        cell.fill        = _header_fill()
        cell.alignment   = _centered()
        cell.border      = _thin_border(COLOR_BORDER)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # ── Données ───────────────────────────────────────────────────────────────
    atts = db.query(Attachement).order_by(
        Attachement.marche_numero,
        Attachement.att_numero
    ).all()

    for row_idx, att in enumerate(atts, start=4):
        ws.row_dimensions[row_idx].height = 22
        fill = _row_fill(row_idx)

        # Calcule montant total et nbre articles
        total_montant = sum(
            (a.montant_total or 0) for a in att.articles
        )
        nbre_articles = len(att.articles)

        # Badge source
        source_label = {
            "claude":          "🔵 Claude",
            "tesseract+claude":"🟢 Hybride",
            "tesseract":       "⚫ Tesseract",
            "manuel":          "✏️ Manuel",
            "error":           "❌ Erreur",
        }.get(att.source or "manuel", att.source or "Manuel")

        row_data = [
            att.entreprise    or "",
            att.date_document or "",
            att.marche_numero or "",
            att.marche_nom    or "",
            att.date_debut    or "",
            att.date_fin      or "",
            att.att_numero,
            nbre_articles     if nbre_articles > 0 else "",
            f"{total_montant:,.2f}" if total_montant > 0 else "",
            source_label,
            att.pdf_path      or "",
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = fill
            cell.border    = _thin_border(COLOR_BORDER)
            cell.font      = _normal_font()
            cell.alignment = _left_aligned() if col_idx in (1, 4, 11) else _centered()

            # Lien cliquable pour le PDF
            if col_idx == 11 and att.pdf_path:
                cell.hyperlink = f"file:///{att.pdf_path.replace(chr(92), '/')}"
                cell.font      = _link_font()

    # ── Freeze pane ──────────────────────────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Filtre automatique ────────────────────────────────────────────────────
    if atts:
        last_row = 3 + len(atts)
        ws.auto_filter.ref = f"A3:K{last_row}"

    # ── Onglet Articles (détail) ──────────────────────────────────────────────
    _write_articles_sheet(wb, db, atts)

    wb.save(EXCEL_PATH)
    print(f"[EXCEL] Sauvegardé : {EXCEL_PATH} ({len(atts)} attachements)")


def _write_articles_sheet(wb, db, atts):
    """Crée un onglet 'Articles' avec toutes les lignes d'articles."""
    if "Articles" in wb.sheetnames:
        del wb["Articles"]

    ws = wb.create_sheet("Articles", 1)

    # Titre
    ws.merge_cells("A1:G1")
    tc = ws["A1"]
    tc.value     = "📋 DÉTAIL DES ARTICLES — TOUS ATTACHEMENTS"
    tc.font      = Font(bold=True, size=13, color=COLOR_HEADER_FG, name="Calibri")
    tc.fill      = PatternFill("solid", fgColor=COLOR_ACCENT)
    tc.alignment = _centered()
    ws.row_dimensions[1].height = 28

    # En-têtes
    headers = [
        ("Marché N°", 14), ("Att N°", 9), ("Entreprise", 22),
        ("Article",   40), ("Quantité", 12), ("Unité", 12),
        ("Prix Unit. DH", 16), ("Montant DH", 16),
    ]
    ws.row_dimensions[2].height = 26
    for col_idx, (h, w) in enumerate(headers, start=1):
        cell           = ws.cell(row=2, column=col_idx, value=h)
        cell.font      = _header_font()
        cell.fill      = _header_fill()
        cell.alignment = _centered()
        cell.border    = _thin_border(COLOR_BORDER)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    row_idx = 3
    for att in atts:
        for art in att.articles:
            fill = _row_fill(row_idx)
            mt   = art.montant_total or (
                (art.quantite or 0) * (art.prix_unitaire or 0)
            )
            row_data = [
                att.marche_numero or "",
                att.att_numero,
                att.entreprise or "",
                art.article    or "",
                art.quantite,
                art.unite      or "",
                f"{art.prix_unitaire:,.2f}" if art.prix_unitaire else "",
                f"{mt:,.2f}"               if mt               else "",
            ]
            ws.row_dimensions[row_idx].height = 20
            for col_idx, value in enumerate(row_data, start=1):
                cell           = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill      = fill
                cell.border    = _thin_border(COLOR_BORDER)
                cell.font      = _normal_font()
                cell.alignment = _left_aligned() if col_idx == 4 else _centered()
            row_idx += 1

    if row_idx > 3:
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:H{row_idx - 1}"
