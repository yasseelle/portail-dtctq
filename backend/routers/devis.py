"""
routers/devis.py  — version corrigée (pdf_path fix)
"""

import os
import re
from pathlib import Path
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, func
from pydantic import BaseModel

import models, auth
from database import get_db

router = APIRouter(prefix="/devis", tags=["Devis"])

SUPER_ADMIN  = "84488R"
STORAGE_DIR  = Path(r"C:\devis\storage")
EXCEL_FILE   = r"C:\devis\etat_devis_2026.xlsx"

MONTHS_FR = {
    1:"Janvier",2:"Février",3:"Mars",4:"Avril",
    5:"Mai",6:"Juin",7:"Juillet",8:"Août",
    9:"Septembre",10:"Octobre",11:"Novembre",12:"Décembre",
}

# =============================================================================
# HELPERS PDF
# =============================================================================

def _resolve_pdf(pdf_path: str) -> Optional[Path]:
    """
    Résout le chemin réel d'un PDF.
    1. Chemin complet direct
    2. Nom de fichier dans STORAGE_DIR
    3. Recherche case-insensitive dans STORAGE_DIR
    4. Recherche par date dans le nom (YYYYMMDD)
    """
    if not pdf_path:
        return None

    # 1. Chemin complet direct
    p = Path(pdf_path)
    if p.exists():
        return p

    # 2. Nom de fichier dans STORAGE_DIR
    filename = os.path.basename(pdf_path)
    if filename:
        candidate = STORAGE_DIR / filename
        if candidate.exists():
            return candidate

        # 3. Case-insensitive
        try:
            for f in STORAGE_DIR.iterdir():
                if f.name.lower() == filename.lower():
                    return f
        except:
            pass

        # 4. Chercher par date dans le nom (YYYYMMDD)
        m = re.search(r"(\d{8})", filename)
        if m:
            date_part = m.group(1)
            try:
                for f in STORAGE_DIR.iterdir():
                    if date_part in f.name:
                        return f
            except:
                pass

    return None


def _pdf_exists(pdf_path: str) -> bool:
    return _resolve_pdf(pdf_path) is not None


def _pdf_filename(pdf_path: str) -> str:
    """Retourne le nom réel du fichier PDF."""
    resolved = _resolve_pdf(pdf_path)
    if resolved:
        return resolved.name
    return os.path.basename(pdf_path) if pdf_path else ""


# =============================================================================
# AUTH
# =============================================================================

def is_super_admin(user: models.User) -> bool:
    return user.matricule == SUPER_ADMIN

def require_super_admin(current_user=Depends(auth.require_admin)):
    if not is_super_admin(current_user):
        raise HTTPException(403, "Réservé au super admin uniquement")
    return current_user


# =============================================================================
# SCHEMAS
# =============================================================================

class DevisCreate(BaseModel):
    reference:    str
    destinataire: str
    objet:        str
    montant_ttc:  Optional[str] = ""
    date_devis:   Optional[str] = ""
    mois:         Optional[str] = ""

class DevisUpdate(BaseModel):
    reference:    Optional[str] = None
    destinataire: Optional[str] = None
    objet:        Optional[str] = None
    montant_ttc:  Optional[str] = None
    date_devis:   Optional[str] = None
    mois:         Optional[str] = None


# =============================================================================
# LECTURE
# =============================================================================

@router.get("/")
def list_devis(
    search: Optional[str] = Query(None),
    mois:   Optional[str] = Query(None),
    page:   int = Query(1, ge=1),
    limit:  int = Query(20, ge=1, le=100),
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    q = db.query(models.Devis)

    if search:
        term = f"%{search}%"
        q = q.filter(or_(
            models.Devis.reference.ilike(term),
            models.Devis.destinataire.ilike(term),
            models.Devis.objet.ilike(term),
            models.Devis.montant_ttc.ilike(term),
        ))
    if mois:
        q = q.filter(models.Devis.mois.ilike(f"%{mois}%"))

    total = q.count()
    items = q.order_by(models.Devis.id.desc()).offset((page-1)*limit).limit(limit).all()

    return {
        "total": total,
        "page":  page,
        "pages": (total + limit - 1) // limit,
        "items": [{
            "id":           r.id,
            "reference":    r.reference,
            "destinataire": r.destinataire,
            "objet":        r.objet,
            "montant_ttc":  r.montant_ttc,
            "date_devis":   r.date_devis,
            "mois":         r.mois,
            "pdf_filename": _pdf_filename(r.pdf_path),       # ← nom réel
            "has_pdf":      _pdf_exists(r.pdf_path),          # ← robuste
            "created_at":   r.created_at.strftime("%d/%m/%Y %H:%M") if r.created_at else "—",
        } for r in items]
    }


@router.get("/stats")
def get_stats(
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    total = db.query(models.Devis).count()
    monthly  = db.query(models.Devis.mois, func.count(models.Devis.id)).group_by(models.Devis.mois).all()
    top_dest = db.query(models.Devis.destinataire, func.count(models.Devis.id))\
                 .group_by(models.Devis.destinataire)\
                 .order_by(func.count(models.Devis.id).desc()).limit(8).all()

    total_montant = 0.0
    for (m,) in db.query(models.Devis.montant_ttc).all():
        if m:
            clean = re.sub(r"[^\d,.]", "", str(m)).replace(",", ".")
            try: total_montant += float(clean)
            except: pass

    return {
        "total":         total,
        "total_montant": f"{total_montant:,.2f}".replace(",", " ").replace(".", ","),
        "monthly":       [{"mois":m, "count":c} for m,c in monthly],
        "top_dest":      [{"destinataire":d, "count":c} for d,c in top_dest],
    }


@router.get("/pdf/{filename}")
def serve_pdf(
    filename: str,
    current_user=Depends(auth.get_current_user),
):
    # Sécurité
    if ".." in filename or "/" in filename or "\\" in filename:
        raise HTTPException(400, "Nom invalide")

    # Cherche dans STORAGE_DIR
    pdf_path = STORAGE_DIR / filename
    if not pdf_path.exists():
        # Case-insensitive
        for f in STORAGE_DIR.iterdir():
            if f.name.lower() == filename.lower():
                pdf_path = f
                break

    if not pdf_path.exists():
        raise HTTPException(404, f"PDF introuvable : {filename}")

    return FileResponse(str(pdf_path), media_type="application/pdf", filename=filename)


@router.get("/debug")
def debug_pdf_links(
    current_user=Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    """Endpoint de diagnostic — liste l'état des liens PDF."""
    all_devis = db.query(models.Devis).all()
    storage_files = [f.name for f in STORAGE_DIR.iterdir() if f.suffix.lower() == ".pdf"] if STORAGE_DIR.exists() else []

    return {
        "storage_dir":   str(STORAGE_DIR),
        "storage_files": storage_files,
        "devis": [{
            "id":           d.id,
            "reference":    d.reference,
            "pdf_path_raw": d.pdf_path,
            "pdf_filename": _pdf_filename(d.pdf_path),
            "has_pdf":      _pdf_exists(d.pdf_path),
            "resolved":     str(_resolve_pdf(d.pdf_path)) if _resolve_pdf(d.pdf_path) else None,
        } for d in all_devis],
    }


# =============================================================================
# SYNC EXCEL → DB
# =============================================================================

@router.post("/sync")
def sync_excel(
    current_user=Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    import openpyxl
    if not Path(EXCEL_FILE).exists():
        raise HTTPException(404, "Fichier Excel introuvable")

    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    created = updated = skipped = 0

    for sheet_name in wb.sheetnames:
        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[5:]:
            if not any(v for v in row): continue
            ref     = str(row[1]).strip() if row[1] else ""
            dest    = str(row[2]).strip() if row[2] else ""
            objet   = str(row[3]).strip() if row[3] else ""
            montant = str(row[4]).strip() if row[4] else ""
            date_d  = str(row[6]).strip() if len(row)>6 and row[6] else ""
            if not ref and not objet: continue

            existing = db.query(models.Devis).filter(models.Devis.reference == ref).first()
            if existing:
                changed = False
                if existing.objet        != objet:   existing.objet        = objet;   changed = True
                if existing.montant_ttc  != montant: existing.montant_ttc  = montant; changed = True
                if existing.destinataire != dest:    existing.destinataire = dest;    changed = True
                if changed: updated += 1
                else:        skipped += 1
            else:
                db.add(models.Devis(reference=ref, destinataire=dest, objet=objet,
                    montant_ttc=montant, date_devis=date_d, mois=sheet_name))
                created += 1

    db.commit(); wb.close()
    return {"created": created, "updated": updated, "skipped": skipped}


# =============================================================================
# CRUD — SUPER ADMIN
# =============================================================================

@router.post("/")
def create_devis(req: DevisCreate, current_user=Depends(require_super_admin), db: Session=Depends(get_db)):
    if req.reference:
        if db.query(models.Devis).filter(models.Devis.reference == req.reference).first():
            raise HTTPException(400, f"Référence '{req.reference}' déjà enregistrée")
    mois = req.mois or MONTHS_FR.get(datetime.now().month, "")
    record = models.Devis(reference=req.reference, destinataire=req.destinataire,
        objet=req.objet, montant_ttc=req.montant_ttc or "",
        date_devis=req.date_devis or "", mois=mois)
    db.add(record); db.commit(); db.refresh(record)
    return {"message": f"✅ Devis '{req.reference}' créé", "id": record.id}


@router.put("/{devis_id}")
def update_devis(devis_id: int, req: DevisUpdate, current_user=Depends(require_super_admin), db: Session=Depends(get_db)):
    record = db.query(models.Devis).filter(models.Devis.id == devis_id).first()
    if not record: raise HTTPException(404, "Devis introuvable")
    if req.reference    is not None: record.reference    = req.reference
    if req.destinataire is not None: record.destinataire = req.destinataire
    if req.objet        is not None: record.objet        = req.objet
    if req.montant_ttc  is not None: record.montant_ttc  = req.montant_ttc
    if req.date_devis   is not None: record.date_devis   = req.date_devis
    if req.mois         is not None: record.mois         = req.mois
    db.commit()
    return {"message": "✅ Devis mis à jour"}


@router.delete("/{devis_id}")
def delete_devis(devis_id: int, current_user=Depends(require_super_admin), db: Session=Depends(get_db)):
    record = db.query(models.Devis).filter(models.Devis.id == devis_id).first()
    if not record: raise HTTPException(404, "Devis introuvable")
    if record.pdf_path:
        try: Path(record.pdf_path).unlink(missing_ok=True)
        except: pass
    db.delete(record); db.commit()
    return {"message": f"✅ Devis '{record.reference}' supprimé"}