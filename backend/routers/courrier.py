"""
routers/courrier.py — API Courrier, Bordereau, Départ/Réception
+ PDF serving + Excel sync

FIX SYNC : Les lignes supprimées manuellement ne reviennent plus.
Mécanisme : chaque ligne a un hash unique (contenu Excel).
Quand une ligne est supprimée via l'API database manager,
son hash est ajouté à la table `sync_blacklist`.
La sync vérifie la blacklist avant toute insertion.
"""

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, func, text
from typing import Optional
from pathlib import Path
import os
import re
import hashlib

import models, auth
from database import get_db

router = APIRouter(prefix="/courrier", tags=["Courrier"])


def _date_sort_expr(date_col):
    """Convertit DD/MM/YYYY en YYYYMMDD pour un tri SQL correct."""
    return func.substr(date_col, 7, 4) + func.substr(date_col, 4, 2) + func.substr(date_col, 1, 2)

STORAGE_DIRS = {
    "arrivee":   r"C:\courrier\storage",
    "bordereau": r"C:\bordereau_envoi\storage",
    "depart":    r"C:\courrier_depart_reception\depart_storage",
    "reception": r"C:\courrier_depart_reception\reception_storage",
}


# =============================================================================
# ── HELPERS ───────────────────────────────────────────────────────────────────
# =============================================================================

def get_pdf_filename(pdf_path: str) -> str:
    if not pdf_path:
        return ""
    return os.path.basename(pdf_path)


def _parse_date(val):
    if not val: return ""
    s = str(val).strip()
    from datetime import datetime
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try: return datetime.strptime(s, fmt).strftime("%d/%m/%Y")
        except: pass
    return s


def _clean(val):
    if val is None: return ""
    return str(val).strip()


def _make_hash(*parts: str) -> str:
    """
    Crée un hash MD5 court à partir des champs clés d'une ligne Excel.
    Ce hash identifie une ligne de façon unique indépendamment de son ID en DB.
    """
    combined = "|".join(str(p).strip().lower() for p in parts)
    return hashlib.md5(combined.encode("utf-8")).hexdigest()[:16]


def _ensure_blacklist_table(db: Session):
    """
    Crée la table sync_blacklist si elle n'existe pas encore.
    Cette table mémorise les hash des lignes supprimées manuellement.
    """
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS sync_blacklist (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            table_name  TEXT NOT NULL,
            row_hash    TEXT NOT NULL,
            deleted_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(table_name, row_hash)
        )
    """))
    db.commit()


def _is_blacklisted(db: Session, table_name: str, row_hash: str) -> bool:
    """Vérifie si un hash est dans la blacklist (= supprimé manuellement)."""
    result = db.execute(
        text("SELECT 1 FROM sync_blacklist WHERE table_name=:t AND row_hash=:h"),
        {"t": table_name, "h": row_hash}
    ).fetchone()
    return result is not None


def _add_to_blacklist(db: Session, table_name: str, row_hash: str):
    """Ajoute un hash à la blacklist."""
    try:
        db.execute(
            text("INSERT OR IGNORE INTO sync_blacklist (table_name, row_hash) VALUES (:t, :h)"),
            {"t": table_name, "h": row_hash}
        )
        db.commit()
    except Exception:
        pass


def _remove_from_blacklist(db: Session, table_name: str, row_hash: str):
    """Retire un hash de la blacklist (si on veut forcer la réinsertion)."""
    db.execute(
        text("DELETE FROM sync_blacklist WHERE table_name=:t AND row_hash=:h"),
        {"t": table_name, "h": row_hash}
    )
    db.commit()


def _get_blacklist_count(db: Session, table_name: str) -> int:
    result = db.execute(
        text("SELECT COUNT(*) FROM sync_blacklist WHERE table_name=:t"),
        {"t": table_name}
    ).fetchone()
    return result[0] if result else 0


# =============================================================================
# ── ENDPOINT : BLACKLIST UN ENREGISTREMENT (appelé par database manager) ──────
# =============================================================================

@router.post("/blacklist/{table_name}/{row_id}")
def blacklist_row(
    table_name: str,
    row_id: int,
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    """
    Marque un enregistrement comme 'supprimé définitivement'.
    Appelé automatiquement par le database manager avant de supprimer une ligne.
    La sync Excel ne réinsèrera plus cette ligne.
    """
    _ensure_blacklist_table(db)

    # Calcule le hash selon la table
    table_models = {
        "courrier":        models.Courrier,
        "bordereau":       models.Bordereau,
        "courrier_depart": models.CourrierDepart,
    }
    model = table_models.get(table_name)
    if not model:
        return {"message": f"Table {table_name} non gérée par la blacklist"}

    row = db.query(model).filter(model.id == row_id).first()
    if not row:
        raise HTTPException(404, "Enregistrement introuvable")

    row_hash = _compute_row_hash(table_name, row)
    _add_to_blacklist(db, table_name, row_hash)

    return {"message": f"✅ Ligne #{row_id} blacklistée — la sync ne la réinsèrera plus", "hash": row_hash}


@router.post("/blacklist/bulk/{table_name}")
def blacklist_bulk(
    table_name: str,
    ids: list,
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    """Blackliste plusieurs enregistrements en une fois (bulk delete)."""
    _ensure_blacklist_table(db)

    table_models = {
        "courrier":        models.Courrier,
        "bordereau":       models.Bordereau,
        "courrier_depart": models.CourrierDepart,
    }
    model = table_models.get(table_name)
    if not model:
        return {"message": f"Table {table_name} non gérée", "blacklisted": 0}

    count = 0
    for row_id in ids:
        row = db.query(model).filter(model.id == row_id).first()
        if row:
            row_hash = _compute_row_hash(table_name, row)
            _add_to_blacklist(db, table_name, row_hash)
            count += 1

    return {"message": f"✅ {count} ligne(s) blacklistée(s)", "blacklisted": count}


@router.get("/blacklist/stats")
def blacklist_stats(
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    """Statistiques de la blacklist."""
    _ensure_blacklist_table(db)
    return {
        "courrier":        _get_blacklist_count(db, "courrier"),
        "bordereau":       _get_blacklist_count(db, "bordereau"),
        "courrier_depart": _get_blacklist_count(db, "courrier_depart"),
    }


@router.delete("/blacklist/reset/{table_name}")
def reset_blacklist(
    table_name: str,
    current_user=Depends(auth.require_admin),
    db: Session = Depends(get_db),
):
    """Vide la blacklist d'une table (permet de tout resync depuis Excel)."""
    _ensure_blacklist_table(db)
    db.execute(
        text("DELETE FROM sync_blacklist WHERE table_name=:t"),
        {"t": table_name}
    )
    db.commit()
    return {"message": f"✅ Blacklist {table_name} vidée — la prochaine sync réinsèrera toutes les lignes Excel"}


def _compute_row_hash(table_name: str, row) -> str:
    """Calcule le hash d'une ligne existante en DB selon sa table."""
    if table_name == "courrier":
        return _make_hash(
            getattr(row, "expediteur", ""),
            getattr(row, "date_courrier", ""),
            getattr(row, "objet", ""),
            getattr(row, "mois", ""),
        )
    elif table_name == "bordereau":
        return _make_hash(
            getattr(row, "reference", ""),
            getattr(row, "objet", ""),
            getattr(row, "destinataire", ""),
        )
    elif table_name == "courrier_depart":
        return _make_hash(
            getattr(row, "reference", ""),
            getattr(row, "date_depart", ""),
            getattr(row, "objet", ""),
        )
    return _make_hash(str(getattr(row, "id", "")))


# =============================================================================
# ── PDF SERVING ───────────────────────────────────────────────────────────────
# =============================================================================

@router.get("/pdf/{source}/{filename}")
def serve_pdf(
    source: str,
    filename: str,
    current_user=Depends(auth.get_current_user),
):
    if source not in STORAGE_DIRS:
        raise HTTPException(status_code=400, detail="Source invalide")

    if ".." in filename or "/" in filename or "\\" in filename:
        raise HTTPException(status_code=400, detail="Nom de fichier invalide")

    pdf_path = Path(STORAGE_DIRS[source]) / filename

    if not pdf_path.exists():
        storage = Path(STORAGE_DIRS[source])
        if storage.exists():
            for f in storage.iterdir():
                if f.name.lower() == filename.lower():
                    pdf_path = f
                    break

    if not pdf_path.exists():
        raise HTTPException(status_code=404, detail=f"PDF introuvable : {filename}")

    return FileResponse(
        path=str(pdf_path),
        media_type="application/pdf",
        filename=filename,
    )


# =============================================================================
# ── EXCEL SYNC — avec protection blacklist ────────────────────────────────────
# =============================================================================

@router.post("/sync/arrivee")
def sync_arrivee(
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    import openpyxl
    excel_path = r"C:\courrier\etat_courrier_2026.xlsx"
    if not Path(excel_path).exists():
        raise HTTPException(status_code=404, detail="Fichier Excel introuvable")

    _ensure_blacklist_table(db)
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    created = updated = skipped = blacklisted_skip = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[5:]:
            if not any(v for v in row): continue
            expediteur = _clean(row[1]) if len(row) > 1 else ""
            date_val   = _parse_date(row[2]) if len(row) > 2 else ""
            objet      = _clean(row[3]) if len(row) > 3 else ""
            if not expediteur and not objet: continue

            # ── VÉRIFICATION BLACKLIST ──────────────────────────────────────
            row_hash = _make_hash(expediteur, date_val, objet, sheet_name)
            if _is_blacklisted(db, "courrier", row_hash):
                blacklisted_skip += 1
                continue  # ← Cette ligne a été supprimée manuellement, on l'ignore

            existing = db.query(models.Courrier).filter(
                models.Courrier.expediteur    == expediteur,
                models.Courrier.date_courrier == date_val,
                models.Courrier.mois          == sheet_name,
            ).first()

            if existing:
                if existing.objet != objet:
                    existing.objet = objet; updated += 1
                else:
                    skipped += 1
            else:
                db.add(models.Courrier(
                    expediteur=expediteur, date_courrier=date_val,
                    objet=objet, mois=sheet_name,
                ))
                created += 1

    db.commit(); wb.close()
    return {
        "created":           created,
        "updated":           updated,
        "skipped":           skipped,
        "blacklisted_skip":  blacklisted_skip,  # lignes ignorées car supprimées manuellement
    }


@router.post("/sync/bordereau")
def sync_bordereau(
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    import openpyxl
    excel_path = r"C:\bordereau_envoi\etat_bordereaux_envoi_2026.xlsx"
    if not Path(excel_path).exists():
        raise HTTPException(status_code=404, detail="Fichier Excel introuvable")

    _ensure_blacklist_table(db)
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    created = updated = skipped = blacklisted_skip = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[5:]:
            if not any(v for v in row): continue
            reference    = _clean(row[1]) if len(row) > 1 else ""
            destinataire = _clean(row[3]) if len(row) > 3 else ""
            objet        = _clean(row[4]) if len(row) > 4 else ""
            if not reference and not objet: continue

            # ── VÉRIFICATION BLACKLIST ──────────────────────────────────────
            row_hash = _make_hash(reference, objet, destinataire)
            if _is_blacklisted(db, "bordereau", row_hash):
                blacklisted_skip += 1
                continue

            existing = db.query(models.Bordereau).filter(
                models.Bordereau.reference == reference,
            ).first()

            if existing:
                if existing.objet != objet or existing.destinataire != destinataire:
                    existing.objet = objet; existing.destinataire = destinataire; updated += 1
                else:
                    skipped += 1
            else:
                db.add(models.Bordereau(reference=reference, destinataire=destinataire, objet=objet))
                created += 1

    db.commit(); wb.close()
    return {
        "created":          created,
        "updated":          updated,
        "skipped":          skipped,
        "blacklisted_skip": blacklisted_skip,
    }


@router.post("/sync/depart")
def sync_depart(
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    import openpyxl
    excel_path = r"C:\courrier_depart_reception\etat_courrier_2026.xlsx"
    if not Path(excel_path).exists():
        raise HTTPException(status_code=404, detail="Fichier Excel introuvable")

    _ensure_blacklist_table(db)
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    created = updated = skipped = blacklisted_skip = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[5:]:
            if not any(v for v in row): continue
            reference      = _clean(row[1]) if len(row) > 1 else ""
            date_depart    = _parse_date(row[2]) if len(row) > 2 else ""
            destinataire   = _clean(row[3]) if len(row) > 3 else ""
            objet          = _clean(row[4]) if len(row) > 4 else ""
            date_reception = _parse_date(row[7]) if len(row) > 7 else ""
            if not reference and not objet: continue

            # ── VÉRIFICATION BLACKLIST ──────────────────────────────────────
            row_hash = _make_hash(reference, date_depart, objet)
            if _is_blacklisted(db, "courrier_depart", row_hash):
                blacklisted_skip += 1
                continue

            existing = db.query(models.CourrierDepart).filter(
                models.CourrierDepart.reference == reference,
            ).first()

            if existing:
                changed = False
                if existing.objet != objet: existing.objet = objet; changed = True
                if existing.date_reception != date_reception: existing.date_reception = date_reception; changed = True
                if changed: updated += 1
                else:        skipped += 1
            else:
                db.add(models.CourrierDepart(
                    reference=reference, date_depart=date_depart,
                    destinataire=destinataire, objet=objet,
                    date_reception=date_reception, mois=sheet_name,
                ))
                created += 1

    db.commit(); wb.close()
    return {
        "created":          created,
        "updated":          updated,
        "skipped":          skipped,
        "blacklisted_skip": blacklisted_skip,
    }


@router.post("/sync/all")
def sync_all(
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    results = {}
    for name, fn in [("arrivee", sync_arrivee), ("bordereau", sync_bordereau), ("depart", sync_depart)]:
        try:
            results[name] = fn(current_user=current_user, db=db)
        except Exception as e:
            results[name] = {"error": str(e)}
    return results


# =============================================================================
# ── COURRIER ARRIVÉE ──────────────────────────────────────────────────────────
# =============================================================================

@router.get("/arrivee")
def get_courrier(
    search:   Optional[str] = Query(None),
    mois:     Optional[str] = Query(None),
    page:     int = Query(1, ge=1),
    limit:    int = Query(20, ge=1, le=100),
    sort_by:  str = Query("date", regex="^(date|expediteur|objet|mois)$"),
    sort_dir: str = Query("desc", regex="^(asc|desc)$"),
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    q = db.query(models.Courrier)
    if search:
        term = f"%{search}%"
        q = q.filter(or_(
            models.Courrier.expediteur.ilike(term),
            models.Courrier.objet.ilike(term),
            models.Courrier.date_courrier.ilike(term),
        ))
    if mois:
        q = q.filter(models.Courrier.mois.ilike(f"%{mois}%"))

    sort_map = {
        "date":       _date_sort_expr(models.Courrier.date_courrier),
        "expediteur": models.Courrier.expediteur,
        "objet":      models.Courrier.objet,
        "mois":       models.Courrier.mois,
    }
    sort_expr = sort_map.get(sort_by, _date_sort_expr(models.Courrier.date_courrier))
    sort_expr = sort_expr.desc() if sort_dir == "desc" else sort_expr.asc()

    total = q.count()
    items = q.order_by(sort_expr).offset((page-1)*limit).limit(limit).all()

    return {
        "total": total, "page": page,
        "pages": (total + limit - 1) // limit,
        "items": [{
            "id":            r.id,
            "expediteur":    r.expediteur,
            "date_courrier": r.date_courrier,
            "objet":         r.objet,
            "mois":          r.mois,
            "pdf_path":      r.pdf_path,
            "pdf_filename":  get_pdf_filename(r.pdf_path),
        } for r in items]
    }


@router.get("/arrivee/stats")
def get_courrier_stats(
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    total   = db.query(models.Courrier).count()
    monthly = db.query(models.Courrier.mois, func.count(models.Courrier.id))\
                .group_by(models.Courrier.mois).all()
    top_exp = db.query(models.Courrier.expediteur, func.count(models.Courrier.id))\
                .group_by(models.Courrier.expediteur)\
                .order_by(func.count(models.Courrier.id).desc()).limit(10).all()
    return {
        "total":   total,
        "monthly": [{"mois": m, "count": c} for m, c in monthly],
        "top_expediteurs": [{"expediteur": e, "count": c} for e, c in top_exp],
    }


# =============================================================================
# ── BORDEREAU ─────────────────────────────────────────────────────────────────
# =============================================================================

@router.get("/bordereau")
def get_bordereau(
    search:   Optional[str] = Query(None),
    page:     int = Query(1, ge=1),
    limit:    int = Query(20, ge=1, le=100),
    sort_by:  str = Query("date", regex="^(date|reference|destinataire|objet)$"),
    sort_dir: str = Query("desc", regex="^(asc|desc)$"),
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    q = db.query(models.Bordereau)
    if search:
        term = f"%{search}%"
        q = q.filter(or_(
            models.Bordereau.reference.ilike(term),
            models.Bordereau.destinataire.ilike(term),
            models.Bordereau.objet.ilike(term),
        ))

    sort_map = {
        "date":        models.Bordereau.created_at,
        "reference":   models.Bordereau.reference,
        "destinataire":models.Bordereau.destinataire,
        "objet":       models.Bordereau.objet,
    }
    sort_expr = sort_map.get(sort_by, models.Bordereau.created_at)
    sort_expr = sort_expr.desc() if sort_dir == "desc" else sort_expr.asc()

    total = q.count()
    items = q.order_by(sort_expr).offset((page-1)*limit).limit(limit).all()
    return {
        "total": total, "page": page,
        "pages": (total + limit - 1) // limit,
        "items": [{
            "id":           r.id,
            "reference":    r.reference,
            "destinataire": r.destinataire,
            "objet":        r.objet,
            "pdf_path":     r.pdf_path,
            "pdf_filename": get_pdf_filename(r.pdf_path),
            "created_at":   r.created_at.strftime("%d/%m/%Y") if r.created_at else "—",
        } for r in items]
    }


@router.get("/bordereau/stats")
def get_bordereau_stats(
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    total    = db.query(models.Bordereau).count()
    top_dest = db.query(models.Bordereau.destinataire, func.count(models.Bordereau.id))\
                 .group_by(models.Bordereau.destinataire)\
                 .order_by(func.count(models.Bordereau.id).desc()).limit(10).all()
    return {
        "total": total,
        "top_destinataires": [{"destinataire": d, "count": c} for d, c in top_dest],
    }


# =============================================================================
# ── DÉPART / RÉCEPTION ────────────────────────────────────────────────────────
# =============================================================================

@router.get("/depart")
def get_depart(
    search: Optional[str] = Query(None),
    mois:   Optional[str] = Query(None),
    page:   int = Query(1, ge=1),
    limit:  int = Query(20, ge=1, le=100),
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    q = db.query(models.CourrierDepart)
    if search:
        term = f"%{search}%"
        q = q.filter(or_(
            models.CourrierDepart.reference.ilike(term),
            models.CourrierDepart.destinataire.ilike(term),
            models.CourrierDepart.objet.ilike(term),
        ))
    if mois:
        q = q.filter(models.CourrierDepart.mois.ilike(f"%{mois}%"))

    sort_expr = _date_sort_expr(models.CourrierDepart.date_depart).desc()

    total = q.count()
    items = q.order_by(sort_expr).offset((page-1)*limit).limit(limit).all()
    return {
        "total": total, "page": page,
        "pages": (total + limit - 1) // limit,
        "items": [{
            "id":               r.id,
            "reference":        r.reference,
            "date_depart":      r.date_depart,
            "destinataire":     r.destinataire,
            "objet":            r.objet,
            "date_reception":   r.date_reception,
            "has_reception":    bool(r.date_reception),
            "mois":             r.mois,
            "pdf_filename":     get_pdf_filename(r.pdf_depart_path),
        } for r in items]
    }


@router.get("/depart/stats")
def get_depart_stats(
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    total      = db.query(models.CourrierDepart).count()
    with_recep = db.query(models.CourrierDepart)\
                   .filter(models.CourrierDepart.date_reception != "").count()
    top_dest   = db.query(models.CourrierDepart.destinataire, func.count(models.CourrierDepart.id))\
                   .group_by(models.CourrierDepart.destinataire)\
                   .order_by(func.count(models.CourrierDepart.id).desc()).limit(8).all()
    return {
        "total": total, "with_recep": with_recep, "pending": total - with_recep,
        "top_destinataires": [{"destinataire": d, "count": c} for d, c in top_dest],
    }


# =============================================================================
# ── STATS GLOBALES ────────────────────────────────────────────────────────────
# =============================================================================

@router.get("/stats/global")
def get_global_stats(
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    return {
        "courrier_total":  db.query(models.Courrier).count(),
        "bordereau_total": db.query(models.Bordereau).count(),
        "depart_total":    db.query(models.CourrierDepart).count(),
        "docs_rh_total":   db.query(models.DocumentRH).count(),
    }